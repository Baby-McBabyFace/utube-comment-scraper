import re
import pytz
from datetime import datetime
from html import unescape
from googleapiclient.discovery import build
from openpyxl import Workbook

api_key = 'YOUR_API_KEY'  # YOUR API KEY HERE
youtube = build('youtube', 'v3', developerKey=api_key)

def convert_UTC_to_tz(utc_time_str, target_timezone="Asia/Seoul"):
    utc_time = datetime.fromisoformat(utc_time_str.replace("Z", "+00:00"))
    target_time = utc_time.astimezone(pytz.timezone(target_timezone))
    return target_time.replace(tzinfo=None)

def get_video_id(youtube_url):
    return youtube_url.split("v=")[-1]

def get_video_title(video_id):
    response = youtube.videos().list(part="snippet", id=video_id).execute()
    return response["items"][0]["snippet"]["title"] if response["items"] else "Unknown Title"

def clean_comment_text(text):
    text_no_br_tags = re.sub(r'<br>', '\n', text)
    text_no_html_tags = re.sub(r'<.*?>', '', text_no_br_tags)
    return unescape(text_no_html_tags) 

def fetch_replies(parent_id):
    replies = []
    next_page_token = None

    replyCount = 1
    while True:
        request = youtube.comments().list(
            part="snippet",
            parentId=parent_id,
            maxResults=100,
            pageToken=next_page_token
        )
        response = request.execute()

        for item in response.get("items", []):
            reply_data = item["snippet"]
            reply = {
                "count" : replyCount,
                "username": reply_data["authorDisplayName"],
                "reply_date": reply_data["publishedAt"],
                "reply_text": clean_comment_text(reply_data["textDisplay"]),
                "like_count": reply_data["likeCount"]
            }
            replies.append(reply)
            replyCount += 1
        next_page_token = response.get("nextPageToken")
        if not next_page_token:
            break

    return replies

def fetch_comments(video_url, max_results=100, order="relevance"):
    video_id = get_video_id(video_url)
    comments = []
    next_page_token = None

    commentCount = 1

    if order == "time":
        while True:
            request = youtube.commentThreads().list(
                part="snippet",
                videoId=video_id,
                maxResults=100,
                pageToken=next_page_token,
                order=order
            )
            response = request.execute()

            for item in response.get("items", []):
                
                comment_data = item["snippet"]["topLevelComment"]["snippet"]
                comment = {
                    "count" : commentCount,
                    "username": comment_data["authorDisplayName"],
                    "comment_date": comment_data["publishedAt"],
                    "comment_text": clean_comment_text(comment_data["textDisplay"]),
                    "like_count": comment_data["likeCount"],
                    "reply_count": item["snippet"]["totalReplyCount"],
                    "replies": []
                }

                total_reply_count = item["snippet"]["totalReplyCount"]
                if total_reply_count > 0:
                    comment["replies"] = fetch_replies(item["id"])

                comments.append(comment)
                commentCount += 1

            next_page_token = response.get("nextPageToken")
            if not next_page_token:
                break

        return comments[-max_results:] if len(comments) > max_results else comments

    else:
        while len(comments) < max_results:
            request = youtube.commentThreads().list(
                part="snippet",
                videoId=video_id,
                maxResults=min(100, max_results - len(comments)),
                pageToken=next_page_token,
                order=order
            )
            response = request.execute()

            for item in response.get("items", []):
                comment_data = item["snippet"]["topLevelComment"]["snippet"]
                comment = {
                    "count" : commentCount,
                    "username": comment_data["authorDisplayName"],
                    "comment_date": comment_data["publishedAt"],
                    "comment_text": clean_comment_text(comment_data["textDisplay"]),
                    "like_count": comment_data["likeCount"],
                    "reply_count": item["snippet"]["totalReplyCount"],
                    "replies": []
                }

                total_reply_count = item["snippet"]["totalReplyCount"]
                if total_reply_count > 0:
                    comment["replies"] = fetch_replies(item["id"])

                comments.append(comment)
                commentCount += 1

            next_page_token = response.get("nextPageToken")
            if not next_page_token:
                break

        return comments[:max_results]

def get_video_details(video_url):
    video_id = get_video_id(video_url)  # Ensure you have a function to extract the video ID
    request = youtube.videos().list(
        part="snippet,contentDetails,statistics,localizations",
        id=video_id
    )
    response = request.execute()

    if not response['items']:
        return None 

    video_info = response['items'][0]

    title_original = video_info['snippet']['title']
    title_english = video_info['localizations'].get('en', {}).get('title', 'N/A')
    date_uploaded = video_info['snippet']['publishedAt']
    video_link = f"https://www.youtube.com/watch?v={video_id}"
    duration = video_info['contentDetails']['duration']
    description = video_info['snippet']['description']
    view_count = video_info['statistics'].get('viewCount', 0)
    like_count = video_info['statistics'].get('likeCount', 0)
    comment_count = video_info['statistics'].get('commentCount', 0)


    # Compile all information into a dictionary
    video_details = {
        "title_english": title_english,
        "title_original": title_original,
        "date_uploaded": date_uploaded,
        "date_uploaded_KST": convert_UTC_to_tz(utc_time_str = date_uploaded),
        "video_link": video_link,
        "duration": duration,
        "description": description,
        "view_count": view_count,
        "like_count": like_count,
        "comment_count": comment_count
    }

    # Save to Excel

    return video_details, title_original

def save_video_details_to_excel(wb, video_details):
    ws = wb.create_sheet(title="Video Details")

    # Add headers
    for col_num, key in enumerate(video_details.keys(), start=1):
        ws.cell(row=1, column=col_num, value=key)

    # Add values
    for col_num, value in enumerate(video_details.values(), start=1):
        ws.cell(row=2, column=col_num, value=value)

def save_comments_to_excel(wb, comments):
    ws = wb.create_sheet(title= "Video Comments")

    # Header row for comments
    ws.append([
        "Comment No.", "Top-Level Comment Username", "Comment Date (UTC)", "Comment Date (KST)", "Comment Text", "Comment Likes", "Number of Replies",
        "Reply No.", "Reply Username", "Reply Date (UTC)", "Reply Date (KST)", "Reply Text", "Reply Likes"
    ])

    for comment in comments:
        top_level_comment = [
            comment["count"],
            comment["username"],
            comment["comment_date"],
            convert_UTC_to_tz(utc_time_str=comment["comment_date"]),
            comment["comment_text"],
            comment["like_count"],
            comment["reply_count"]
        ]
        if comment["replies"]:
            for reply in comment["replies"]:
                ws.append(top_level_comment + [
                    reply["count"],
                    reply["username"],
                    reply["reply_date"],
                    convert_UTC_to_tz(utc_time_str=reply["reply_date"]),
                    reply["reply_text"],
                    reply["like_count"]
                ])
        else:
            ws.append(top_level_comment + ["", "", "", "", ""])
            
def save_all_data(video_details, comments, title):
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Save video details and comments in separate sheets
    save_video_details_to_excel(wb, video_details)
    save_comments_to_excel(wb, comments)

    filename = f"results/{sanitize_sheet_title(title)} scraped on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}.xlsx"
    wb.save(filename)
    print(f"Saved all data to {filename}")

def sanitize_sheet_title(title):
    # Replace invalid characters with an underscore
    return re.sub(r'[\\/:*?"<>|]', '_', title)

def main():
    with open('links.txt', 'r') as links:
        for link in links:
            video_url = link.rstrip('\n')
            video_id = get_video_id(video_url)
            video_title = get_video_title(video_id)
            comments = fetch_comments(video_url, max_results=100, order="relevance") # Change order to relevance OR time
            video_details, title = get_video_details(video_url)
            save_all_data(video_details=video_details, comments=comments, title=title)
            
if __name__ == "__main__":
    main()